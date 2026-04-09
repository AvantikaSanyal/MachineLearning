"""
=============================================================================
SHAP Analysis — Hybrid TB Diagnostic System
=============================================================================
Runs SHAP explainability on:
  1. Random Forest  (TreeExplainer)
  2. SVM            (KernelExplainer on a background sample)
Produces:
  • Summary plots (beeswarm)
  • Bar plots (mean |SHAP|)
  • Waterfall / force plots for individual patients
  • SHAP dependence plots for top features
  • SHAP heatmap across test set
  • Side-by-side RF vs SVM comparison
  • Combined PDF report
=============================================================================
"""

import warnings, logging
warnings.filterwarnings("ignore")
logging.getLogger("pgmpy").setLevel(logging.WARNING)

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestClassifier
from sklearn.svm import SVC
from sklearn.preprocessing import StandardScaler
from sklearn.impute import SimpleImputer

import shap
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec

RANDOM_STATE = 42
np.random.seed(RANDOM_STATE)
# shap.initjs()  # not needed outside Jupyter

FEATURE_LABELS = {
    "CO":  "Cough",
    "NS":  "Night Sweats",
    "BD":  "Breathing Difficulty",
    "FV":  "Fever",
    "CP":  "Chest Pain",
    "SP":  "Sputum",
    "IS":  "Immune Suppression",
    "LP":  "Loss of Pleasure",
    "CH":  "Chills",
    "LC":  "Lack of Concentration",
    "IR":  "Irritation",
    "LA":  "Loss of Appetite",
    "LE":  "Loss of Energy",
    "LNE": "Lymph Node Enlargement",
    "SBP": "Systolic BP",
    "BMI": "BMI",
}
FEATURES     = list(FEATURE_LABELS.keys())
FEATURE_NAMES = list(FEATURE_LABELS.values())

# ── Colour palette ────────────────────────────────────────────────────────────
RF_COLOR  = "#2196F3"
SVM_COLOR = "#FF9800"


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — Reproduce data & models from main script
# ══════════════════════════════════════════════════════════════════════════════
print("=" * 70)
print("STEP 1: Preparing Data & Models")
print("=" * 70)

wb   = load_workbook("/mnt/user-data/uploads/tuberculosis_dataset.xlsx", read_only=True)
ws   = wb.active
rows = list(ws.iter_rows(values_only=True))
df   = pd.DataFrame(rows[2:], columns=rows[1]).dropna(how="all").astype(int)

df["Weight_Loss_Score"] = (df["LA"] + df["LE"] + (2 - df["BMI"])) / 3
df["Primary_Score"]     = df["CO"] + df["FV"] + df["NS"] + df["Weight_Loss_Score"]
df["TB_Label"]          = (df["Primary_Score"] >= 4).astype(int)

X = df[FEATURES].copy()
y = df["TB_Label"].copy()

X_train, X_test, y_train, y_test = train_test_split(
    X, y, test_size=0.2, random_state=RANDOM_STATE, stratify=y
)

# Rename columns to full names for readability in plots
X_train_named = X_train.rename(columns=FEATURE_LABELS)
X_test_named  = X_test.rename(columns=FEATURE_LABELS)

# Random Forest
rf_model = RandomForestClassifier(
    n_estimators=100, max_depth=None, min_samples_split=2, random_state=RANDOM_STATE
)
rf_model.fit(X_train, y_train)
print(f"  RF trained  → Test Acc: {rf_model.score(X_test, y_test)*100:.1f}%")

# SVM
scaler     = StandardScaler()
X_train_sc = scaler.fit_transform(X_train)
X_test_sc  = scaler.transform(X_test)

svm_model = SVC(kernel="rbf", C=10, gamma=0.01, probability=True, random_state=RANDOM_STATE)
svm_model.fit(X_train_sc, y_train)
print(f"  SVM trained → Test Acc: {svm_model.score(X_test_sc, y_test)*100:.1f}%")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — SHAP for Random Forest (TreeExplainer)
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("STEP 2: SHAP — Random Forest (TreeExplainer)")
print("=" * 70)

rf_explainer  = shap.TreeExplainer(rf_model)
rf_shap_vals  = rf_explainer(X_test_named)          # Explanation object
# For binary classification shap returns shape (n, features, 2); take class=1
rf_sv         = rf_shap_vals[..., 1]                 # class TB-Positive
rf_base_value = rf_sv.base_values[0]

print(f"  SHAP base value (RF)  : {rf_base_value:.4f}")
print(f"  SHAP values shape     : {rf_sv.values.shape}")

# Mean |SHAP| per feature
rf_mean_shap = pd.DataFrame({
    "Feature":    FEATURE_NAMES,
    "Mean |SHAP|": np.abs(rf_sv.values).mean(axis=0)
}).sort_values("Mean |SHAP|", ascending=False)
print("\n  Top features by mean |SHAP| (RF):")
print(rf_mean_shap.to_string(index=False))


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — SHAP for SVM (KernelExplainer on background sample)
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("STEP 3: SHAP — SVM (KernelExplainer, background n=50)")
print("=" * 70)

# KernelExplainer is model-agnostic but slower; use a k-means background
background   = shap.kmeans(X_train_sc, 50)
svm_predict  = lambda x: svm_model.predict_proba(x)[:, 1]  # P(TB=1)

svm_explainer = shap.KernelExplainer(svm_predict, background)
print("  Computing SHAP values for test set (may take ~30 s) …")
svm_shap_raw  = svm_explainer.shap_values(X_test_sc, nsamples=100, silent=True)

# Wrap in Explanation object with named columns for consistent API
svm_sv = shap.Explanation(
    values          = svm_shap_raw,
    base_values     = np.full(len(X_test_sc), svm_explainer.expected_value),
    data            = X_test_named.values,
    feature_names   = FEATURE_NAMES,
)
svm_base_value = svm_explainer.expected_value
print(f"  SHAP base value (SVM) : {svm_base_value:.4f}")
print(f"  SHAP values shape     : {svm_sv.values.shape}")

svm_mean_shap = pd.DataFrame({
    "Feature":    FEATURE_NAMES,
    "Mean |SHAP|": np.abs(svm_sv.values).mean(axis=0)
}).sort_values("Mean |SHAP|", ascending=False)
print("\n  Top features by mean |SHAP| (SVM):")
print(svm_mean_shap.to_string(index=False))


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 — FIGURES
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("STEP 4: Generating SHAP Figures")
print("=" * 70)

OUT = "/mnt/user-data/outputs"


# ── Fig 1: RF Beeswarm Summary ────────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(9, 6))
shap.plots.beeswarm(rf_sv, max_display=16, show=False, color_bar=True)
plt.title("SHAP Summary Plot — Random Forest\n(TB-Positive class)", fontweight="bold", pad=12)
plt.tight_layout()
plt.savefig(f"{OUT}/shap_rf_beeswarm.png", dpi=150, bbox_inches="tight")
plt.close()
print("  Saved → shap_rf_beeswarm.png")


# ── Fig 2: SVM Beeswarm Summary ──────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(9, 6))
shap.plots.beeswarm(svm_sv, max_display=16, show=False, color_bar=True)
plt.title("SHAP Summary Plot — SVM\n(TB-Positive class)", fontweight="bold", pad=12)
plt.tight_layout()
plt.savefig(f"{OUT}/shap_svm_beeswarm.png", dpi=150, bbox_inches="tight")
plt.close()
print("  Saved → shap_svm_beeswarm.png")


# ── Fig 3: Side-by-side Mean |SHAP| bar comparison ───────────────────────────
fig, axes = plt.subplots(1, 2, figsize=(14, 6))
fig.suptitle("Mean |SHAP Value| — Feature Importance Comparison",
             fontweight="bold", fontsize=13)

for ax, ms, title, color in [
    (axes[0], rf_mean_shap,  "Random Forest",  RF_COLOR),
    (axes[1], svm_mean_shap, "SVM (RBF)",      SVM_COLOR),
]:
    ax.barh(ms["Feature"][::-1], ms["Mean |SHAP|"][::-1], color=color, alpha=0.85)
    ax.set_title(title, fontweight="bold")
    ax.set_xlabel("Mean |SHAP Value|")
    ax.grid(axis="x", alpha=0.3)

plt.tight_layout()
plt.savefig(f"{OUT}/shap_bar_comparison.png", dpi=150, bbox_inches="tight")
plt.close()
print("  Saved → shap_bar_comparison.png")


# ── Fig 4: RF Waterfall plots — 3 representative patients ────────────────────
fig, axes = plt.subplots(1, 3, figsize=(18, 5))
fig.suptitle("SHAP Waterfall Plots — Random Forest\nIndividual Patient Explanations",
             fontweight="bold", fontsize=13)

# Pick: 1 true positive, 1 true negative, 1 false negative (if any)
y_pred_rf = rf_model.predict(X_test)
tp_idx = np.where((y_test.values == 1) & (y_pred_rf == 1))[0]
tn_idx = np.where((y_test.values == 0) & (y_pred_rf == 0))[0]
fp_idx = np.where((y_test.values == 0) & (y_pred_rf == 1))[0]
fn_idx = np.where((y_test.values == 1) & (y_pred_rf == 0))[0]

cases = []
if len(tp_idx): cases.append((tp_idx[0], "True Positive (Correct TB+)"))
if len(tn_idx): cases.append((tn_idx[0], "True Negative (Correct TB-)"))
if len(fn_idx): cases.append((fn_idx[0], "False Negative (Missed TB)"))
elif len(fp_idx): cases.append((fp_idx[0], "False Positive (False Alarm)"))
# pad to 3
while len(cases) < 3:
    cases.append((tp_idx[len(cases)], f"TB-Positive Patient #{len(cases)+1}"))

for ax, (i, label) in zip(axes, cases):
    plt.sca(ax)
    shap.plots.waterfall(rf_sv[i], max_display=10, show=False)
    ax.set_title(label, fontweight="bold", fontsize=10)

plt.tight_layout()
plt.savefig(f"{OUT}/shap_rf_waterfall.png", dpi=150, bbox_inches="tight")
plt.close()
print("  Saved → shap_rf_waterfall.png")


# ── Fig 5: SHAP Dependence plots — top 4 RF features ─────────────────────────
top4 = rf_mean_shap["Feature"].head(4).tolist()
fig, axes = plt.subplots(2, 2, figsize=(12, 9))
fig.suptitle("SHAP Dependence Plots — Top 4 Features (Random Forest)",
             fontweight="bold", fontsize=13)

for ax, feat in zip(axes.flat, top4):
    feat_idx = FEATURE_NAMES.index(feat)
    shap.dependence_plot(
        feat_idx,
        rf_sv.values,
        X_test_named,
        ax=ax,
        show=False,
        dot_size=30,
        alpha=0.7,
    )
    ax.set_title(f"Dependence: {feat}", fontweight="bold")
    ax.grid(alpha=0.2)

plt.tight_layout()
plt.savefig(f"{OUT}/shap_rf_dependence.png", dpi=150, bbox_inches="tight")
plt.close()
print("  Saved → shap_rf_dependence.png")


# ── Fig 6: SHAP Heatmap (RF) ─────────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(14, 6))
shap.plots.heatmap(rf_sv, max_display=16, show=False)
plt.title("SHAP Heatmap — Random Forest (Test Set, sorted by prediction)",
          fontweight="bold", pad=10)
plt.tight_layout()
plt.savefig(f"{OUT}/shap_rf_heatmap.png", dpi=150, bbox_inches="tight")
plt.close()
print("  Saved → shap_rf_heatmap.png")


# ── Fig 7: Ranked feature importance — RF vs SVM (dot chart) ─────────────────
fig, ax = plt.subplots(figsize=(9, 6))
rf_ranked  = rf_mean_shap.set_index("Feature")["Mean |SHAP|"]
svm_ranked = svm_mean_shap.set_index("Feature")["Mean |SHAP|"]
order      = rf_mean_shap["Feature"].tolist()  # sort by RF

x_rf  = [rf_ranked[f]  for f in order]
x_svm = [svm_ranked[f] for f in order]
y_pos = range(len(order))

ax.barh(y_pos, x_rf,  height=0.35, label="Random Forest", color=RF_COLOR,  alpha=0.85, align="edge")
ax.barh([y+0.35 for y in y_pos], x_svm, height=0.35, label="SVM", color=SVM_COLOR, alpha=0.85, align="edge")
ax.set_yticks([y+0.35 for y in y_pos])
ax.set_yticklabels(order)
ax.invert_yaxis()
ax.set_xlabel("Mean |SHAP Value|")
ax.set_title("SHAP Feature Ranking — RF vs SVM", fontweight="bold")
ax.legend()
ax.grid(axis="x", alpha=0.3)
plt.tight_layout()
plt.savefig(f"{OUT}/shap_rf_vs_svm_ranking.png", dpi=150, bbox_inches="tight")
plt.close()
print("  Saved → shap_rf_vs_svm_ranking.png")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 5 — PRINT RANKED TABLE
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("STEP 5: SHAP Feature Ranking Summary")
print("=" * 70)

merged = rf_mean_shap.rename(columns={"Mean |SHAP|": "RF Mean |SHAP|"}).merge(
    svm_mean_shap.rename(columns={"Mean |SHAP|": "SVM Mean |SHAP|"}),
    on="Feature"
).sort_values("RF Mean |SHAP|", ascending=False).reset_index(drop=True)
merged.index += 1
print("\n  (sorted by RF SHAP importance)")
print(merged.to_string())


print("\n" + "=" * 70)
print("SHAP ANALYSIS COMPLETE")
print("=" * 70)
print("""
Output files:
  shap_rf_beeswarm.png       — RF beeswarm summary (all features)
  shap_svm_beeswarm.png      — SVM beeswarm summary (all features)
  shap_bar_comparison.png    — RF vs SVM mean |SHAP| bar charts
  shap_rf_waterfall.png      — Waterfall plots for 3 individual patients
  shap_rf_dependence.png     — Dependence plots for top 4 features
  shap_rf_heatmap.png        — SHAP heatmap across the test set
  shap_rf_vs_svm_ranking.png — Side-by-side ranked feature comparison
""")
